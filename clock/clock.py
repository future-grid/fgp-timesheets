import sys
import openpyxl

timesheet   = sys.argv[1]
hours       = sys.argv[2]
project     = sys.argv[3]
today       = sys.argv[4]

try:
    # bash doesnt handle strings with spaces.
    description = sys.argv[5].replace("~", " ")
except Exception as ex:
    description = False # set the description to not write
    print("No description found, not writing one.")

wb    = openpyxl.load_workbook(timesheet)
sheet = wb[wb.sheetnames[0]]

# get day as row in excel sheet
for i in range(3, 10):
    day = sheet.cell(row=i, column=1).value
    if day == today:
        row = i

# write data
sheet.cell(row=row, column=3).value = project
sheet.cell(row=row, column=4).value = hours
if not description is False:
    sheet.cell(row=row, column=5).value = description

# save sheet
wb.save(timesheet)

print("Wrote to timesheet.")
