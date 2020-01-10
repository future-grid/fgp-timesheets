#!/usr/local/bin/python3

import openpyxl as op
import time, datetime, os, csv, json, sys

# function will return an object with the date of the friday, and the initials of the creator
# "filepath" is expected to be of form "TimeSheet_wYYYYMMDD_FL" FL = initials
def get_data_from_title(filepath):
    title = filepath[len("timesheets/TimeSheet_w"):-len(".xlsx")]
    initials = title[-2:]
    date = datetime.datetime.strptime(title[:-3], '%Y%m%d')
    date = date.timetuple()
    date = time.strftime('%d-%m-%y', date)

    return {"initials": initials, "date": date}

# returns a full day with arbitrary many projects / hours. uses the friday date to determine the day's date
def get_day_from_data(row,friday_date):
    # this array defines the difference between the given day and friday. eg friday-friday = 0, friday-wednesday=2
    week_order = ["Friday", "Thursday", "Wednesday", "Tuesday", "Monday", "Sunday"]
    friday_date = datetime.datetime.strptime(friday_date, '%d-%m-%y')
    date = friday_date - datetime.timedelta(days=week_order.index(row[0])) # row[0] is always the day of the week eg. "Tuesday"
    date = date.timetuple()
    date = time.strftime('%d-%m-%y', date)

    projects=[]
    for i in range(2, len(row)-1, 3):
        if row[i] == None:
            break
        projects.append({"project":row[i], "hours":float(row[i+1]), "description":row[i+2]})

    return {"date":date, "projects":projects}

# returns a list of the entire worksheet using very cool python stuff
def iter_rows(ws):
    for row in ws.iter_rows():
        yield [cell.value for cell in row]

def read_timesheet(filepath):
    wb = op.load_workbook(filepath)
    ws = wb.active # get the active worksheet (there will be only one worksheet per timesheet)
    if ws.cell(row=4,column=1).value == "=A3+1":
        ws.delete_cols(1)

    titledata = get_data_from_title(filepath)
    sheetdata = list(iter_rows(ws))
    days = []

    for i in range(3,9):
        days.append(get_day_from_data(sheetdata[i], titledata["date"]))

    return {"initials": titledata["initials"], "friday": titledata["date"], "days":days}

def save_sheets(sheet_list):
    try:
        with open('cache.json', 'w') as fpath:
            json.dump({'sheets' : sheet_list}, fpath)
    except OSError as err:
        print("Got an exception!")
        print(err)

def load_sheets(root_path):
    if '--load-cache' in sys.argv:
        print("Loading from cache.json")
        with open('cache.json', "r") as fpath:
            contents = json.load(fpath)
            sheets = contents["sheets"]
    else:
        directory = os.fsencode(root_path)
        sheets = []
        for file in os.listdir(directory):
            filename = os.fsdecode(file)
            if filename.endswith(".xlsx"):
                try:
                    sheets.append(read_timesheet(root_path + "/" + filename))
                except Exception as e:
                    print("Could not read timesheet " + filename + "!")
                    print(e)

    if '--cache' in sys.argv:
        save_sheets(sheets)

    return sheets

# this function will return the earliest date from a list of sheets
def get_date(sheets, looking_for="earliest"):
    best_date = None
    for sheet in sheets:
        for day in sheet["days"]:
            date = datetime.datetime.strptime(day['date'], '%d-%m-%y')
            best_date = date if best_date is None else best_date
            if looking_for is "earliest":
                best_date = date if date < best_date else best_date
            elif looking_for is "latest":
                best_date = date if date > best_date else best_date
    return best_date

def generate_timeline_overview(sheets):
    early_date = get_date(sheets, looking_for="earliest")
    late_date = get_date(sheets, looking_for="latest") + datetime.timedelta(days=1)
    current_date = early_date

    first_row = ["Date", "Employee", "Project", "Hours", "Description"]
    rows = [first_row]

    # iterate over days
    while True:
        date = datetime.datetime.strftime(current_date, '%d-%m-%y')
        # increase day count by one, if we are at end - stop
        current_date = current_date + datetime.timedelta(days=1)
        for sheet in sheets:
            employee = sheet["initials"]
            for day in sheet["days"]:
                if day["date"] == date:
                    projects = []
                    for project in day["projects"]: # iterate over projects on day
                        project_code = project["project"]
                        project_hours = float(project["hours"])
                        project_description = project["description"]
                        projects_concatenated = False

                        for p in projects:
                            if p["project_code"] == project_code:
                               # concat descriptions and hours
                               p["hours"].append(project_hours)
                               p["description"].append(project_description)
                               projects_concatenated = True

                        # if we've already added the project to another one, don't double count
                        if projects_concatenated:
                            continue

                        projects.append({
                            'date' : date,
                            'employee' : employee,
                            'project_code' : project_code,
                            'hours' : [project_hours],
                            'description' : [project_description]
                        })

                    for project in projects:
                        description = ''
                        hours = 0
                        date = project["date"]
                        employee = project["employee"]
                        project_code = project["project_code"]
                        if len(project["description"]) > 1:
                            hour_list = project["hours"]
                            description_list = project["description"]

                            for hour, desc in zip(hour_list, description_list):
                                description += (desc + " ("+str(hour)+"), ")
                                hours += hour
                            description = description[:-2] # remove trailing comma
                        else:
                            description = project["description"][0]
                            hours = project["hours"][0]
                        rows.append ([
                            date, employee, project_code, hours, description
                        ])

        if current_date > late_date:
            break

        # save file as csv
        with open("report.csv", "w", newline="") as f:
            writer = csv.writer(f)
            writer.writerows(rows)

if len(sys.argv[1]) > 0:
    sheets = load_sheets(sys.argv[1])
else:
    sheets = load_sheets("timesheets")
generate_timeline_overview(sheets)
